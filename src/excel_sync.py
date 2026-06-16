"""
Excel同步模块 - 读写配箱表.xlsm
支持数据双向同步、冲突检测、自动备份
"""

import os
import shutil
import time
from typing import Optional, Dict, List, Any, Tuple
from datetime import datetime
from dataclasses import dataclass

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from .engine import PeiXiangEngine, OrderItem, BoxItem, dino_to_hub


class ExcelSync:
    """Excel文件读写同步管理器"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb: Optional[openpyxl.Workbook] = None
        self.last_modified: float = 0
        self.backup_dir = os.path.join(os.path.dirname(filepath), "peixiang_backups")
        self._sheet_cache: Dict[str, Any] = {}

    # ===== 文件操作 =====
    def open(self, read_only: bool = False) -> bool:
        """打开Excel文件"""
        try:
            self.wb = load_workbook(self.filepath, data_only=False, keep_vba=True)
            self.last_modified = os.path.getmtime(self.filepath)
            self._sheet_cache = {}
            return True
        except Exception as e:
            raise RuntimeError(f"打开文件失败: {e}")

    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
            self.wb = None

    def save(self) -> bool:
        """保存文件"""
        try:
            self.wb.save(self.filepath)
            self.last_modified = os.path.getmtime(self.filepath)
            self._sheet_cache = {}
            return True
        except Exception as e:
            raise RuntimeError(f"保存文件失败: {e}")

    def save_as(self, filepath: str) -> bool:
        """另存为"""
        try:
            self.wb.save(filepath)
            return True
        except Exception as e:
            raise RuntimeError(f"另存为失败: {e}")

    def backup(self) -> str:
        """创建自动备份"""
        os.makedirs(self.backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(self.backup_dir, f"配箱表_backup_{timestamp}.xlsm")
        shutil.copy2(self.filepath, backup_path)
        # 保留最近10个备份
        backups = sorted([f for f in os.listdir(self.backup_dir) if f.startswith("配箱表_backup_")])
        for old in backups[:-10]:
            os.remove(os.path.join(self.backup_dir, old))
        return backup_path

    def check_external_modify(self) -> bool:
        """检查文件是否被外部修改"""
        if not os.path.exists(self.filepath):
            return False
        current_mtime = os.path.getmtime(self.filepath)
        return current_mtime > self.last_modified

    # ===== 读取汇总表 =====
    def read_summary_data(self) -> Tuple[List[Dict], Dict[str, str]]:
        """
        读取汇总表数据，构建配箱引擎所需的索引数据
        返回: (箱号数据列表, 牌号映射字典)
        """
        ws = self.wb['汇总']
        rows = []

        # 读取左侧数据区 (H-S列, 从第6行开始)
        for r in range(6, ws.max_row + 1):
            product_desc = ws.cell(r, 8).value   # H: 产品描述
            box_number = ws.cell(r, 15).value     # O: 箱号

            # 跳过空行和"总计"行
            if not product_desc or not box_number:
                continue
            if str(product_desc).strip() == '总计':
                continue

            row_data = {
                'row': r,
                'product_desc': str(product_desc).strip() if product_desc else '',
                'atd': ws.cell(r, 9).value,       # I: 实际离港日期
                'etd': ws.cell(r, 11).value,       # K: 计划离港日期
                'transit_port': ws.cell(r, 12).value,  # L: 中转港
                'dino': str(ws.cell(r, 13).value or '').strip(),     # M: DINO
                'waybill': str(ws.cell(r, 14).value or '').strip(),  # N: 运单号
                'box_number': str(box_number).strip(),
                'weight': ws.cell(r, 16).value,    # P: 重量
                'batch': str(ws.cell(r, 17).value or '').strip(),    # Q: 批次
                'ship_company': str(ws.cell(r, 18).value or '').strip(),  # R: 船公司
                'device': str(ws.cell(r, 19).value or '').strip(),   # S: 装置
                # 右侧区域
                'ae_product': str(ws.cell(r, 31).value or '').strip(),
                'ak_dino': str(ws.cell(r, 37).value or '').strip(),
                'al_weight': ws.cell(r, 38).value,
            }
            rows.append(row_data)

        return rows

    # ===== 读取牌号映射 =====
    def read_material_mapping(self) -> Dict[str, str]:
        """读取配箱公式sheet的A:B牌号映射表"""
        ws = self.wb['配箱公式']
        mapping = {}
        for r in range(2, ws.max_row + 1):
            a = ws.cell(r, 1).value  # A: erp物料名称
            b = ws.cell(r, 2).value  # B: 物流表物料名称
            if a and b and not str(b).startswith('='):
                mapping[str(a).strip()] = str(b).strip()
        return mapping

    # ===== 读取订单数据 =====
    def read_orders(self) -> List[Dict]:
        """读取配箱公式sheet中的订单数据"""
        ws = self.wb['配箱公式']
        orders = []
        for r in range(2, ws.max_row + 1):
            po = ws.cell(r, 26).value  # Z: 客户订单号
            if not po or str(po).strip() == '':
                continue

            order = {
                'row': r,
                'po': str(po).strip(),
                'product_name': str(ws.cell(r, 30).value or '').strip(),  # AD
                'quantity': ws.cell(r, 35).value or 0,                    # AI
                'sales_org': str(ws.cell(r, 29).value or '').strip(),     # AC
                'brand': str(ws.cell(r, 44).value or '').strip(),         # AR
                'spec_batch': str(ws.cell(r, 21).value or '').strip(),    # U
                'other_req': str(ws.cell(r, 22).value or '').strip(),     # V
                'spec_di': str(ws.cell(r, 23).value or '').strip(),       # W
                'incoterm': str(ws.cell(r, 24).value or '').strip(),      # X
                'carrier': str(ws.cell(r, 31).value or '').strip(),       # AE
                'status': str(ws.cell(r, 27).value or '').strip(),        # AA
                'line_no': ws.cell(r, 28).value,                          # AB
            }
            # 转换数量
            try:
                order['quantity'] = float(order['quantity'])
            except (ValueError, TypeError):
                order['quantity'] = 0

            orders.append(order)
        return orders

    # ===== 读取物流跟踪数据 =====
    def read_logistics_data(self) -> Dict[str, Dict]:
        """读取物流跟踪sheet数据，按运单号索引"""
        ws = self.wb['物流跟踪']
        data = {}
        for r in range(2, ws.max_row + 1):
            waybill = ws.cell(r, 20).value  # T: 运单号
            if not waybill or str(waybill).strip() == '':
                continue
            waybill = str(waybill).strip()
            data[waybill] = {
                'row': r,
                'order_type': str(ws.cell(r, 1).value or ''),
                'dino': str(ws.cell(r, 2).value or ''),
                'sono': str(ws.cell(r, 4).value or ''),
                'product_code': str(ws.cell(r, 11).value or ''),
                'product_desc': str(ws.cell(r, 12).value or ''),
                'weight': ws.cell(r, 13).value,
                'box_count': ws.cell(r, 14).value,
                'box_type': str(ws.cell(r, 15).value or ''),
                'batch': str(ws.cell(r, 16).value or ''),
                'box_number': str(ws.cell(r, 17).value or ''),
                'seal_number': str(ws.cell(r, 18).value or ''),
                'ship_company': str(ws.cell(r, 19).value or ''),
                'waybill': waybill,
                'ship_name': str(ws.cell(r, 21).value or ''),
                'etd': ws.cell(r, 22).value,
                'load_date': ws.cell(r, 23).value,
                'atd': ws.cell(r, 24).value,
                'transit_port': str(ws.cell(r, 25).value or ''),
                'second_ship': str(ws.cell(r, 28).value or ''),
                'eta': ws.cell(r, 31).value,
                'ata': ws.cell(r, 32).value,
                'device': str(ws.cell(r, 35).value or ''),
            }
        return data

    # ===== 读取ASN数据 =====
    def read_asn_data(self) -> List[Dict]:
        """读取ASN sheet数据"""
        ws = self.wb['ASN']
        data = []
        for r in range(2, ws.max_row + 1):
            po = ws.cell(r, 1).value
            if not po or str(po).strip() == '':
                continue
            data.append({
                'row': r,
                'po': str(po).strip(),
                'doc_no': str(ws.cell(r, 2).value or ''),
                'status': str(ws.cell(r, 3).value or ''),
                'delivery_method': str(ws.cell(r, 4).value or ''),
                'sales_org': str(ws.cell(r, 5).value or ''),
                'batch': str(ws.cell(r, 6).value or ''),
                'date': ws.cell(r, 7).value,
                'warehouse': str(ws.cell(r, 8).value or ''),
                'line_seq': ws.cell(r, 9).value,
                'material': str(ws.cell(r, 10).value or ''),
                'quantity': ws.cell(r, 11).value,
                'unit': str(ws.cell(r, 12).value or ''),
                'container_no': str(ws.cell(r, 20).value or ''),
                'seal_no': str(ws.cell(r, 21).value or ''),
                'bill_no': str(ws.cell(r, 22).value or ''),
            })
        return data

    # ===== 写入配箱结果 =====
    def write_peixiang_results(self, orders: List[OrderItem]):
        """
        将配箱计算结果写回配箱公式sheet
        只写入公式列(C-S列的计算结果)，保留原始数据列
        """
        ws = self.wb['配箱公式']
        for order in orders:
            r = order.row
            ws.cell(r, 3).value = order.seq1              # C: 序列1
            ws.cell(r, 4).value = order.seq2              # D: 序列2
            ws.cell(r, 5).value = order.query1             # E: 查询1
            ws.cell(r, 6).value = order.seq3               # F: 序列3
            ws.cell(r, 8).value = order.brand_mapped       # H: 牌号
            ws.cell(r, 10).value = order.box_number        # J: 箱号
            ws.cell(r, 13).value = order.device            # M: 装置
            ws.cell(r, 14).value = order.coa               # N: COA
            if order.etd:
                ws.cell(r, 15).value = order.etd           # O: 离港时间
            ws.cell(r, 16).value = order.remaining         # P: 剩余数量
            ws.cell(r, 17).value = order.weight_net        # Q: 净重
            ws.cell(r, 18).value = order.pending           # R: 待离数量
            ws.cell(r, 19).value = order.di                # S: DI

    # ===== 写入配箱表sheet =====
    def write_peixiang_table(self, results: List[Dict]):
        """
        将配箱结果写入配箱表sheet
        results: [{dino, sono, product_code, product_desc, weight, box_count,
                   box_type, batch, box_number, seal, ship_company, waybill, ...}]
        """
        ws = self.wb['配箱表']
        # 从第2行开始写入
        for idx, item in enumerate(results):
            r = idx + 2
            ws.cell(r, 1).value = item.get('order_type', 'HUB')
            ws.cell(r, 2).value = item.get('dino', '')
            ws.cell(r, 4).value = item.get('sono', '')
            ws.cell(r, 6).value = item.get('seller', 'LYB')
            ws.cell(r, 9).value = item.get('origin_port', '')
            ws.cell(r, 10).value = item.get('dest_port', '')
            ws.cell(r, 11).value = item.get('product_code', '')
            ws.cell(r, 12).value = item.get('product_desc', '')
            ws.cell(r, 13).value = item.get('weight', 0)
            ws.cell(r, 14).value = item.get('box_count', 1)
            ws.cell(r, 15).value = item.get('box_type', '')
            ws.cell(r, 16).value = item.get('batch', '')
            ws.cell(r, 17).value = item.get('box_number', '')
            ws.cell(r, 18).value = item.get('seal', '')
            ws.cell(r, 19).value = item.get('ship_company', '')
            ws.cell(r, 20).value = item.get('waybill', '')
            ws.cell(r, 21).value = item.get('ship_name', '')
            ws.cell(r, 35).value = item.get('device', '')

    # ===== 插入ASN数据 =====
    def insert_asn_data(self, asn_rows: List[Dict]):
        """
        在ASN sheet顶部插入行（不覆盖原数据）
        asn_rows: ASN数据行列表
        """
        ws = self.wb['ASN']
        # 插入空行
        insert_count = len(asn_rows)
        ws.insert_rows(2, amount=insert_count)

        # 写入数据
        for idx, row_data in enumerate(asn_rows):
            r = idx + 2
            # ASN列映射
            col_map = {
                1: 'po', 2: 'doc_no', 3: 'status', 4: 'delivery_method',
                5: 'sales_org', 6: 'batch', 7: 'date', 8: 'warehouse',
                9: 'line_seq', 10: 'material', 11: 'quantity', 12: 'unit',
                20: 'container_no', 21: 'seal_no', 22: 'bill_no',
            }
            for col, key in col_map.items():
                val = row_data.get(key, '')
                if val:
                    ws.cell(r, col).value = val

    # ===== 从ERP导出数据导入订单 =====
    def import_orders_from_clipboard(self, order_data: List[List]):
        """
        从剪贴板/CSV数据导入订单到配箱公式sheet
        order_data: 二维列表，每行是一个订单行
        """
        ws = self.wb['配箱公式']
        # 找到第一个空行
        start_row = 2
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 26).value is None or str(ws.cell(r, 26).value).strip() == '':
                start_row = r
                break
        else:
            start_row = ws.max_row + 1

        # 写入订单数据到对应列
        # ERP导出列映射: Z, AA, AB, AC, AD, AE, AF, AG, AH, AI, AJ, AK, AL, AM, AN, AO, AP...
        erp_col_map = {
            0: 26,   # 客户订单号 → Z
            1: 27,   # 单据状态 → AA
            2: 28,   # 行号 → AB
            3: 29,   # 销售组织 → AC
            4: 30,   # 物料名称 → AD
            5: 31,   # LYB陆运承运商 → AE
            6: 32,   # LYB承运商联系人 → AF
            7: 33,   # 规格型号 → AG
            8: 34,   # 计量单位 → AH
            9: 35,   # 数量 → AI
            10: 36,  # 单价 → AJ
            11: 37,  # 金额 → AK
            12: 38,  # 含税单价 → AL
            13: 39,  # 税额 → AM
            14: 40,  # 价税合计 → AN
            15: 41,  # 发货组织 → AO
            16: 42,  # 销售组 → AP
            17: 43,  # 客户订单分录序号 → AQ
            18: 44,  # 牌号 → AR
            19: 45,  # 装船要求 → AS
            20: 46,  # 交货指示号 → AT
            21: 47,  # 贸易条款 → AU
            22: 48,  # 订单来源 → AV
            23: 49,  # 总毛重 → AW
            24: 50,  # 总净重 → AX
        }

        for row_idx, row_data in enumerate(order_data):
            r = start_row + row_idx
            for col_idx, value in enumerate(row_data):
                if col_idx in erp_col_map and value is not None:
                    ws.cell(r, erp_col_map[col_idx]).value = value
            # 设置默认值
            ws.cell(r, 21).value = "."   # U: 配箱指定批次
            ws.cell(r, 22).value = "."   # V: 其他配箱要求
            ws.cell(r, 23).value = "."   # W: 配箱指定DI
            ws.cell(r, 20).value = 0     # T: 累计通知发货数量

        return start_row

    # ===== 生成发货通知单导入数据 =====
    def generate_shipment_notice(self, orders: List[OrderItem]) -> List[Dict]:
        """
        根据配箱结果生成发货通知单数据（用于导入ERP）
        """
        notices = []
        for order in orders:
            if not order.box_number or order.box_number == "#N/A":
                continue
            box_item = self._get_box_details(order.box_number)
            if not box_item:
                continue

            notice = {
                '客户采购单号': order.po,
                '销售组织': order.sales_org,
                '物料名称': order.product_name,
                '数量': order.quantity,
                '计量单位': '吨',
                '行号': order.line_no if hasattr(order, 'line_no') else 1,
                '集装箱号': order.box_number,
                '铅封号': box_item.waybill if box_item else '',
                '批次': box_item.batch if box_item else '',
                '装置': order.device,
            }
            notices.append(notice)
        return notices

    def _get_box_details(self, box_number: str) -> Optional[BoxItem]:
        """获取箱号详细信息"""
        # 这个方法会被外部引擎调用
        return None

    # ===== 获取所有sheet名称 =====
    def get_sheet_names(self) -> List[str]:
        return self.wb.sheetnames if self.wb else []

    # ===== 获取配箱公式sheet的数据范围 =====
    def get_order_data_range(self) -> Tuple[int, int]:
        """返回配箱公式sheet中有数据的行范围"""
        ws = self.wb['配箱公式']
        start = 2
        end = 2
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 26).value:
                end = r
        return (start, end)

    # ===== 重新加载汇总表数据 =====
    def reload_to_engine(self, engine: PeiXiangEngine):
        """从Excel重新加载数据到引擎"""
        # 加载牌号映射
        mapping = self.read_material_mapping()
        engine.load_material_mapping(mapping)

        # 加载汇总数据
        summary_rows = self.read_summary_data()
        engine.load_summary_data(summary_rows)

        # 加载物流跟踪数据
        engine.logistics_data = self.read_logistics_data()

        # 加载ASN数据
        engine.asn_data = self.read_asn_data()
